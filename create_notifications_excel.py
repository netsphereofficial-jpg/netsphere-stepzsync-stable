#!/usr/bin/env python3
"""
Generate Excel spreadsheet documenting all notification types in the app.
This includes race, friend, and chat notifications with complete details.
"""

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

# Create workbook
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Notifications"

# Define styles
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=11)
category_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
category_font = Font(bold=True, size=11)
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Set column headers
headers = [
    "No.",
    "Category",
    "Notification Type",
    "Title",
    "Body/Message",
    "Icon",
    "Trigger Event",
    "Trigger Path",
    "Recipients",
    "Additional Data Fields",
    "Notes"
]

# Write headers
for col_num, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_num)
    cell.value = header
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = thin_border

# Set column widths
column_widths = [5, 12, 25, 25, 50, 6, 40, 40, 30, 35, 40]
for i, width in enumerate(column_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = width

# Notification data
notifications = [
    # RACE NOTIFICATIONS
    {
        "category": "RACE",
        "type": "Race Invitation",
        "title": "Race Invitation 🏃‍♂️",
        "body": "{inviterName} invited you to join \"{raceTitle}\"",
        "icon": "🏃‍♂️",
        "trigger": "Document created in race_invites collection with type='received' and isJoinRequest=false",
        "trigger_path": "functions/notifications/triggers/raceTriggers.js - onRaceInviteCreated (lines 40-114)",
        "recipients": "Invited user (toUserId)",
        "data_fields": "type: InviteRace, category: Race, raceId, raceName, inviterUserId, inviterName, startTime (optional), distance (optional), location (optional)",
        "notes": "Sent when race organizer invites someone to join a race. Only processes 'received' type invites to avoid duplicates."
    },
    {
        "category": "RACE",
        "type": "New Join Request",
        "title": "New Join Request 🙋‍♂️",
        "body": "{requesterName} wants to join \"{raceTitle}\"",
        "icon": "🙋‍♂️",
        "trigger": "Document created in race_invites collection with type='received' and isJoinRequest=true",
        "trigger_path": "functions/notifications/triggers/raceTriggers.js - onRaceInviteCreated (lines 40-114)",
        "recipients": "Race organizer (toUserId)",
        "data_fields": "type: NewJoinRequest, category: Race, raceId, raceName, requesterUserId, requesterName, requestedAt",
        "notes": "Sent when user requests to join a race. Organizer receives this notification."
    },
    {
        "category": "RACE",
        "type": "Race Started",
        "title": "Race Started! 🚀",
        "body": "\"{raceTitle}\" has begun! Good luck!",
        "icon": "🚀",
        "trigger": "Race document updated with statusId changed to 3 (ACTIVE)",
        "trigger_path": "functions/notifications/triggers/raceTriggers.js - onRaceStatusChanged (lines 122-218)",
        "recipients": "All race participants",
        "data_fields": "type: RaceBegin, category: Race, raceId, raceName, participantCount (optional), startedAt",
        "notes": "Sent to all participants when race begins (status changes from pending to active)."
    },
    {
        "category": "RACE",
        "type": "Race Completed (Winner - 1st)",
        "title": "Congratulations! 🥇",
        "body": "You won \"{raceTitle}\"! Amazing performance!",
        "icon": "🏆",
        "trigger": "Race document updated with statusId changed to 4 (COMPLETED)",
        "trigger_path": "functions/notifications/triggers/raceTriggers.js - onRaceStatusChanged (lines 122-218)",
        "recipients": "Participant who finished 1st",
        "data_fields": "type: RaceWon, category: Achievement, raceId, raceName, rank: 1, xpEarned (optional), distanceCovered (optional), avgSpeed (optional), completedAt",
        "notes": "Special winner notification for 1st place finisher."
    },
    {
        "category": "RACE",
        "type": "Race Completed (2nd Place)",
        "title": "Great Job! 🥈",
        "body": "You finished 2nd in \"{raceTitle}\"! Well done!",
        "icon": "🥈",
        "trigger": "Race document updated with statusId changed to 4 (COMPLETED)",
        "trigger_path": "functions/notifications/triggers/raceTriggers.js - onRaceStatusChanged (lines 122-218)",
        "recipients": "Participant who finished 2nd",
        "data_fields": "type: RaceCompleted, category: Achievement, raceId, raceName, rank: 2, xpEarned (optional), distanceCovered (optional), avgSpeed (optional), completedAt",
        "notes": "Sent to 2nd place finisher with silver medal emoji."
    },
    {
        "category": "RACE",
        "type": "Race Completed (3rd Place)",
        "title": "Excellent! 🥉",
        "body": "You finished 3rd in \"{raceTitle}\"! Great effort!",
        "icon": "🥉",
        "trigger": "Race document updated with statusId changed to 4 (COMPLETED)",
        "trigger_path": "functions/notifications/triggers/raceTriggers.js - onRaceStatusChanged (lines 122-218)",
        "recipients": "Participant who finished 3rd",
        "data_fields": "type: RaceCompleted, category: Achievement, raceId, raceName, rank: 3, xpEarned (optional), distanceCovered (optional), avgSpeed (optional), completedAt",
        "notes": "Sent to 3rd place finisher with bronze medal emoji."
    },
    {
        "category": "RACE",
        "type": "Race Completed (Other)",
        "title": "Race Completed! 🏃‍♂️",
        "body": "You finished \"{raceTitle}\" in {rank} place!",
        "icon": "🏃‍♂️",
        "trigger": "Race document updated with statusId changed to 4 (COMPLETED)",
        "trigger_path": "functions/notifications/triggers/raceTriggers.js - onRaceStatusChanged (lines 122-218)",
        "recipients": "Participants who finished 4th or lower",
        "data_fields": "type: RaceCompleted, category: Race, raceId, raceName, rank, xpEarned (optional), distanceCovered (optional), avgSpeed (optional), completedAt",
        "notes": "Sent to all other finishers with their rank (4th, 5th, etc.)."
    },
    {
        "category": "RACE",
        "type": "First Finisher",
        "title": "🏁 First to Finish!",
        "body": "Amazing! You're the first to complete \"{raceTitle}\"!",
        "icon": "🏁",
        "trigger": "Race document updated with statusId changed to 6 (ENDING) - first participant crosses finish",
        "trigger_path": "functions/notifications/triggers/raceTriggers.js - onRaceStatusChanged (lines 154-200)",
        "recipients": "First finisher (firstFinisherUserId)",
        "data_fields": "type: RaceFirstFinisher, category: Achievement, raceId, raceName, finishedAt",
        "notes": "Sent when first participant completes the race, triggering the deadline countdown."
    },
    {
        "category": "RACE",
        "type": "Deadline Alert",
        "title": "⏰ Deadline Approaching!",
        "body": "{firstFinisherName} finished first! You have {deadlineMinutes} minutes to complete the race!",
        "icon": "⏰",
        "trigger": "Race document updated with statusId changed to 6 (ENDING) - deadline countdown starts",
        "trigger_path": "functions/notifications/triggers/raceTriggers.js - onRaceStatusChanged (lines 154-200)",
        "recipients": "All active participants who haven't finished yet",
        "data_fields": "type: RaceDeadlineAlert, category: Race, raceId, raceName, firstFinisherName, deadlineMinutes, deadline (ISO timestamp), timestamp",
        "notes": "Sent to remaining active participants when first person finishes, creates urgency to complete."
    },
    {
        "category": "RACE",
        "type": "Race Cancelled",
        "title": "❌ Race Cancelled",
        "body": "The race \"{raceTitle}\" has been cancelled. {reason}",
        "icon": "❌",
        "trigger": "Race document updated with statusId changed to 7 (CANCELLED)",
        "trigger_path": "functions/notifications/triggers/raceTriggers.js - onRaceStatusChanged (lines 202-211)",
        "recipients": "All race participants",
        "data_fields": "type: RaceCancelled, category: Race, raceId, raceName, cancellationReason, cancelledAt",
        "notes": "Sent when race organizer cancels the race. Includes cancellation reason if provided."
    },
    {
        "category": "RACE",
        "type": "Join Request Accepted",
        "title": "Join Request Accepted ✅",
        "body": "{organizerName} accepted your request to join \"{raceTitle}\"",
        "icon": "✅",
        "trigger": "race_invites document updated with status='accepted' and isJoinRequest=true",
        "trigger_path": "functions/notifications/triggers/raceTriggers.js - onRaceInviteAccepted (lines 226-300)",
        "recipients": "User who requested to join (toUserId)",
        "data_fields": "type: JoinRequestAccepted, category: Race, raceId, raceName, organizerUserId, organizerName, acceptedAt",
        "notes": "Sent when organizer approves a join request."
    },
    {
        "category": "RACE",
        "type": "Invite Accepted",
        "title": "Race Invite Accepted 🎉",
        "body": "{accepterName} accepted your invite to \"{raceTitle}\"",
        "icon": "🎉",
        "trigger": "race_invites document updated with status='accepted' and isJoinRequest=false",
        "trigger_path": "functions/notifications/triggers/raceTriggers.js - onRaceInviteAccepted (lines 226-300)",
        "recipients": "Race organizer who sent invite (fromUserId)",
        "data_fields": "type: InviteAccepted, category: Race, raceId, raceName, accepterUserId, accepterName, acceptedAt",
        "notes": "Sent to organizer when invited user accepts race invitation."
    },
    {
        "category": "RACE",
        "type": "Join Request Declined",
        "title": "Join Request Declined",
        "body": "{organizerName} declined your request to join \"{raceTitle}\"",
        "icon": "❌",
        "trigger": "race_invites document updated with status='declined' and isJoinRequest=true",
        "trigger_path": "functions/notifications/triggers/raceTriggers.js - onRaceInviteDeclined (lines 308-382)",
        "recipients": "User who requested to join (toUserId)",
        "data_fields": "type: JoinRequestDeclined, category: Race, raceId, raceName, organizerUserId, organizerName, declinedAt",
        "notes": "Sent when organizer rejects a join request."
    },
    {
        "category": "RACE",
        "type": "Invite Declined",
        "title": "Race Invite Declined",
        "body": "{declinerName} declined your invite to \"{raceTitle}\"",
        "icon": "😔",
        "trigger": "race_invites document updated with status='declined' and isJoinRequest=false",
        "trigger_path": "functions/notifications/triggers/raceTriggers.js - onRaceInviteDeclined (lines 308-382)",
        "recipients": "Race organizer who sent invite (fromUserId)",
        "data_fields": "type: InviteDeclined, category: Race, raceId, raceName, declinerUserId, declinerName, declinedAt",
        "notes": "Sent to organizer when invited user declines race invitation."
    },
    {
        "category": "RACE",
        "type": "Race Created (Confirmation)",
        "title": "Race Created Successfully! 🎉",
        "body": "Your {raceType} \"{raceTitle}\" is ready! Distance: {distance}km.",
        "icon": "🎉",
        "trigger": "New document created in races collection",
        "trigger_path": "functions/notifications/triggers/raceTriggers.js - onRaceCreated (lines 399-443)",
        "recipients": "Race creator (createdBy)",
        "data_fields": "type: RaceCreated, category: Race, raceId, raceName, raceType, distance, scheduledTime (optional), participantCount (optional), createdAt",
        "notes": "Confirmation sent to race creator immediately after race creation."
    },
    {
        "category": "RACE",
        "type": "Public Race Announcement",
        "title": "New Public Race Available! 🏁",
        "body": "{organizerName} created \"{raceTitle}\" - {distance}km. Join now!",
        "icon": "🏁",
        "trigger": "New document created in races collection with raceTypeId=3 (Public)",
        "trigger_path": "functions/notifications/triggers/raceTriggers.js - onRaceCreated (lines 399-443)",
        "recipients": "ALL app users (except creator)",
        "data_fields": "type: PublicRaceAnnouncement, category: Race, raceId, raceName, raceTypeId: 3, distance, organizerName, location (optional), scheduledTime (optional), participantLimit (optional), createdAt",
        "notes": "BROADCAST notification sent to all users when public race is created. Sent in batches of 500 users."
    },
    {
        "category": "RACE",
        "type": "Participant Joined",
        "title": "Someone Joined Your Race! 🎉",
        "body": "{participantName} joined \"{raceTitle}\"",
        "icon": "🎉",
        "trigger": "Manual trigger (not implemented in current code)",
        "trigger_path": "functions/notifications/senders/raceNotifications.js - sendParticipantJoinedNotification (lines 759-792)",
        "recipients": "Race organizer",
        "data_fields": "type: RaceParticipantJoined, category: Race, raceId, raceName, participantId, participantName, joinedAt",
        "notes": "Function exists but trigger not implemented yet. Would notify organizer when someone joins their race."
    },
    {
        "category": "RACE",
        "type": "Overtaking (Overtaker)",
        "title": "Great Overtake! 🚀",
        "body": "Awesome! You overtook {overtakenName} and moved to rank #{newRank}!",
        "icon": "🚀",
        "trigger": "Manual trigger (not implemented in current code)",
        "trigger_path": "functions/notifications/senders/raceNotifications.js - sendOvertakingNotifications (lines 798-903)",
        "recipients": "Participant who overtook",
        "data_fields": "type: RaceOvertaking, category: Achievement, raceId, raceName, newRank, oldRank, overtakenUser, timestamp",
        "notes": "Function exists but trigger not implemented. Would notify when user improves rank."
    },
    {
        "category": "RACE",
        "type": "Overtaking (Overtaken)",
        "title": "You Were Overtaken! ⚡",
        "body": "{overtakerName} just overtook you! Speed up to reclaim your position!",
        "icon": "⚡",
        "trigger": "Manual trigger (not implemented in current code)",
        "trigger_path": "functions/notifications/senders/raceNotifications.js - sendOvertakingNotifications (lines 798-903)",
        "recipients": "Participant who was overtaken",
        "data_fields": "type: RaceOvertaken, category: Race, raceId, raceName, overtakerName, yourNewRank, timestamp",
        "notes": "Function exists but trigger not implemented. Creates competitive pressure."
    },
    {
        "category": "RACE",
        "type": "Overtaking (General Alert)",
        "title": "Overtaking Alert! 🏃‍♂️",
        "body": "{overtakerName} overtook {overtakenName} and moved to rank #{newRank}!",
        "icon": "🏃‍♂️",
        "trigger": "Manual trigger (not implemented in current code)",
        "trigger_path": "functions/notifications/senders/raceNotifications.js - sendOvertakingNotifications (lines 798-903)",
        "recipients": "All other race participants (except overtaker and overtaken)",
        "data_fields": "type: RaceOvertakingGeneral, category: Race, raceId, raceName, overtakingUser, overtakenUser, newRank, timestamp",
        "notes": "Function exists but trigger not implemented. Keeps all participants informed of race dynamics."
    },
    {
        "category": "RACE",
        "type": "Leader Change",
        "title": "New Leader! 👑",
        "body": "{newLeaderName} took the lead in \"{raceTitle}\"!",
        "icon": "👑",
        "trigger": "Manual trigger (not implemented in current code)",
        "trigger_path": "functions/notifications/senders/raceNotifications.js - sendLeaderChangeNotification (lines 909-959)",
        "recipients": "All race participants (except new leader)",
        "data_fields": "type: RaceLeaderChange, category: Race, raceId, raceName, newLeaderUserId, newLeaderName, timestamp",
        "notes": "Function exists but trigger not implemented. Would notify when someone takes 1st place."
    },
    {
        "category": "RACE",
        "type": "Personal Milestone",
        "title": "Milestone Reached! {icon}",
        "body": "Great job! You've completed {milestone}% of \"{raceTitle}\"!",
        "icon": "🎯/⚡/🔥",
        "trigger": "Manual trigger (not implemented in current code)",
        "trigger_path": "functions/notifications/senders/raceNotifications.js - sendMilestonePersonalNotification (lines 1121-1156)",
        "recipients": "Participant who reached milestone",
        "data_fields": "type: RaceMilestonePersonal, category: Achievement, raceId, raceName, milestone (25/50/75), achievedAt",
        "notes": "Function exists but trigger not implemented. Icons: 25%=🎯, 50%=⚡, 75%=🔥. Personal achievement notification."
    },
    {
        "category": "RACE",
        "type": "Milestone Alert",
        "title": "{userName} Hit {milestone}%! {icon}",
        "body": "{userName} reached {milestone}% of \"{raceTitle}\". Keep pushing!",
        "icon": "🎯/⚡/🔥",
        "trigger": "Manual trigger (not implemented in current code)",
        "trigger_path": "functions/notifications/senders/raceNotifications.js - sendMilestoneAlertNotification (lines 1162-1226)",
        "recipients": "All other race participants (except achiever)",
        "data_fields": "type: RaceMilestoneAlert, category: Race, raceId, raceName, achieverName, achieverUserId, milestone (25/50/75), timestamp",
        "notes": "Function exists but trigger not implemented. Informs others when someone hits 25%, 50%, or 75%."
    },
    {
        "category": "RACE",
        "type": "Proximity Alert",
        "title": "🔥 Opponent Approaching!",
        "body": "{chaserName} is only {distanceGap}m behind you! Speed up!",
        "icon": "🔥",
        "trigger": "Manual trigger (not implemented in current code)",
        "trigger_path": "functions/notifications/senders/raceNotifications.js - sendProximityAlertNotification (lines 1232-1273)",
        "recipients": "Participant being chased",
        "data_fields": "type: RaceProximityAlert, category: Race, raceId, raceName, chaserName, chaserUserId, distanceGap (meters), timestamp",
        "notes": "Function exists but trigger not implemented. Would alert when opponent gets within 20m."
    },
    {
        "category": "RACE",
        "type": "Countdown Timer",
        "title": "⏰ {minutesLeft} Minutes Left!",
        "body": "Time is running out in \"{raceTitle}\"! Sprint to the finish!",
        "icon": "⏰",
        "trigger": "Manual trigger (not implemented in current code)",
        "trigger_path": "functions/notifications/senders/raceNotifications.js - sendCountdownTimerNotification (lines 1279-1333)",
        "recipients": "All active participants who haven't finished",
        "data_fields": "type: RaceCountdownTimer, category: Race, raceId, raceName, minutesLeft, timestamp",
        "notes": "Function exists but trigger not implemented. Would send when deadline approaching (e.g., 5 minutes left)."
    },
    {
        "category": "RACE",
        "type": "Race Reminder (15 min)",
        "title": "Race Starting Soon! ⏰",
        "body": "\"{raceTitle}\" starts in 15 minutes. Get ready!",
        "icon": "⏰",
        "trigger": "Scheduled function (not implemented in current code)",
        "trigger_path": "functions/notifications/senders/raceNotifications.js - sendRaceReminder (lines 214-268)",
        "recipients": "All race participants",
        "data_fields": "type: RaceReminder, category: Race, raceId, raceName, reminderType: 15min, startTime (optional), reminderSentAt",
        "notes": "Function exists but trigger not implemented. Would be triggered 15 minutes before race start."
    },
    {
        "category": "RACE",
        "type": "Race Reminder (1 hour)",
        "title": "Race Reminder 🕐",
        "body": "\"{raceTitle}\" starts in 1 hour. Don't forget!",
        "icon": "⏰",
        "trigger": "Scheduled function (not implemented in current code)",
        "trigger_path": "functions/notifications/senders/raceNotifications.js - sendRaceReminder (lines 214-268)",
        "recipients": "All race participants",
        "data_fields": "type: RaceReminder, category: Race, raceId, raceName, reminderType: 1hour, startTime (optional), reminderSentAt",
        "notes": "Function exists but trigger not implemented. Would be triggered 1 hour before race start."
    },
    {
        "category": "RACE",
        "type": "Race Reminder (1 day)",
        "title": "Race Tomorrow 📅",
        "body": "\"{raceTitle}\" is scheduled for tomorrow.",
        "icon": "⏰",
        "trigger": "Scheduled function (not implemented in current code)",
        "trigger_path": "functions/notifications/senders/raceNotifications.js - sendRaceReminder (lines 214-268)",
        "recipients": "All race participants",
        "data_fields": "type: RaceReminder, category: Race, raceId, raceName, reminderType: 1day, startTime (optional), reminderSentAt",
        "notes": "Function exists but trigger not implemented. Would be triggered 1 day before race start."
    },

    # FRIEND/SOCIAL NOTIFICATIONS
    {
        "category": "SOCIAL",
        "type": "Friend Request",
        "title": "New Friend Request 👥",
        "body": "{senderName} wants to be your friend!",
        "icon": "👥",
        "trigger": "Document created in friend_requests collection",
        "trigger_path": "functions/notifications/triggers/friendTriggers.js - onFriendRequestCreated (lines 30-73)",
        "recipients": "Receiver of friend request",
        "data_fields": "type: FriendRequest, category: Social, userId (sender), userName, thumbnail (profile pic, optional), mutualFriends (optional), requestSentAt",
        "notes": "Sent when user sends a friend request to another user."
    },
    {
        "category": "SOCIAL",
        "type": "Friend Request Accepted",
        "title": "Friend Request Accepted! 🎉",
        "body": "{friendName} accepted your friend request!",
        "icon": "🎉",
        "trigger": "friend_requests document updated with status='accepted'",
        "trigger_path": "functions/notifications/triggers/friendTriggers.js - onFriendRequestAccepted (lines 81-126)",
        "recipients": "Original friend request sender",
        "data_fields": "type: FriendAccepted, category: Social, userId (accepter), userName, thumbnail (profile pic, optional), acceptedAt",
        "notes": "Sent to requester when receiver accepts the friend request."
    },
    {
        "category": "SOCIAL",
        "type": "Friend Request Declined",
        "title": "Friend Request Declined 😔",
        "body": "{friendName} declined your friend request.",
        "icon": "😔",
        "trigger": "friend_requests document updated with status='declined'",
        "trigger_path": "functions/notifications/triggers/friendTriggers.js - onFriendRequestDeclined (lines 134-178)",
        "recipients": "Original friend request sender",
        "data_fields": "type: FriendDeclined, category: Social, userId (decliner), userName, declinedAt",
        "notes": "Sent to requester when receiver declines the friend request."
    },
    {
        "category": "SOCIAL",
        "type": "Friend Removed",
        "title": "Friendship Ended 💔",
        "body": "{removerName} removed you from their friends list.",
        "icon": "💔",
        "trigger": "Document deleted from friends collection",
        "trigger_path": "functions/notifications/triggers/friendTriggers.js - onFriendRemoved (lines 190-228)",
        "recipients": "Friend who was removed",
        "data_fields": "type: FriendRemoved, category: Social, userId (remover), userName, removedAt",
        "notes": "Sent when user removes another user from friends list. Only sent to friendId user to avoid duplicate notifications."
    },

    # CHAT NOTIFICATIONS
    {
        "category": "CHAT",
        "type": "Direct Chat Message",
        "title": "New Message from {senderName} 💬",
        "body": "{messageText}",
        "icon": "💬",
        "trigger": "Document created in chat_messages collection",
        "trigger_path": "functions/notifications/triggers/chatTriggers.js - onChatMessageCreated (lines 26-53)",
        "recipients": "Message receiver (receiverId)",
        "data_fields": "type: ChatMessage, category: Chat, senderName, chatRoomId, clickAction: OPEN_CHAT",
        "notes": "Sent for 1-on-1 direct messages. Message truncated to 100 chars in notification."
    },
    {
        "category": "CHAT",
        "type": "Race Chat Message",
        "title": "{raceTitle} 🏃",
        "body": "{senderName}: {messageText}",
        "icon": "🏃",
        "trigger": "Document created in race_chat_messages collection",
        "trigger_path": "functions/notifications/triggers/chatTriggers.js - onRaceChatMessageCreated (lines 61-119)",
        "recipients": "All race chat participants (except sender)",
        "data_fields": "type: RaceChatMessage, category: RaceChat, senderName, raceTitle, raceId, raceChatId, clickAction: OPEN_RACE_CHAT",
        "notes": "Sent to all race chat participants when someone sends a message. Message truncated to 80 chars. Notifications sent in parallel to all participants."
    },
]

# Write data rows
row_num = 2
for idx, notif in enumerate(notifications, 1):
    # Write data
    ws.cell(row=row_num, column=1).value = idx
    ws.cell(row=row_num, column=2).value = notif["category"]
    ws.cell(row=row_num, column=3).value = notif["type"]
    ws.cell(row=row_num, column=4).value = notif["title"]
    ws.cell(row=row_num, column=5).value = notif["body"]
    ws.cell(row=row_num, column=6).value = notif["icon"]
    ws.cell(row=row_num, column=7).value = notif["trigger"]
    ws.cell(row=row_num, column=8).value = notif["trigger_path"]
    ws.cell(row=row_num, column=9).value = notif["recipients"]
    ws.cell(row=row_num, column=10).value = notif["data_fields"]
    ws.cell(row=row_num, column=11).value = notif["notes"]

    # Apply styling
    for col in range(1, 12):
        cell = ws.cell(row=row_num, column=col)
        cell.alignment = Alignment(vertical='top', wrap_text=True)
        cell.border = thin_border

        # Highlight category cells
        if col == 2:
            cell.font = Font(bold=True)

        # Center number column
        if col == 1:
            cell.alignment = Alignment(horizontal='center', vertical='top')

    row_num += 1

# Set row heights
ws.row_dimensions[1].height = 30
for row in range(2, row_num):
    ws.row_dimensions[row].height = 60

# Freeze header row
ws.freeze_panes = 'A2'

# Add summary sheet
summary_ws = wb.create_sheet("Summary", 0)

# Summary header
summary_ws.merge_cells('A1:B1')
summary_ws['A1'] = "Notification System Summary"
summary_ws['A1'].font = Font(bold=True, size=14, color="FFFFFF")
summary_ws['A1'].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
summary_ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
summary_ws.row_dimensions[1].height = 25

# Summary data
summary_data = [
    ["Generated On:", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
    ["Total Notifications:", len(notifications)],
    ["", ""],
    ["Category Breakdown:", ""],
    ["Race Notifications:", len([n for n in notifications if n["category"] == "RACE"])],
    ["Social Notifications:", len([n for n in notifications if n["category"] == "SOCIAL"])],
    ["Chat Notifications:", len([n for n in notifications if n["category"] == "CHAT"])],
    ["", ""],
    ["Implementation Status:", ""],
    ["Active (Triggered):", len([n for n in notifications if "not implemented" not in n["trigger"].lower()])],
    ["Pending (Not Triggered):", len([n for n in notifications if "not implemented" in n["trigger"].lower()])],
    ["", ""],
    ["Key Firestore Triggers:", ""],
    ["race_invites (onCreate)", "Race invitations & join requests"],
    ["races (onUpdate)", "Race status changes (started, completed, cancelled, ending)"],
    ["races (onCreate)", "Race creation & public announcements"],
    ["race_invites (onUpdate)", "Invite/join request acceptance/decline"],
    ["friend_requests (onCreate)", "Friend requests"],
    ["friend_requests (onUpdate)", "Friend request responses"],
    ["friends (onDelete)", "Friend removal"],
    ["chat_messages (onCreate)", "Direct chat messages"],
    ["race_chat_messages (onCreate)", "Race chat messages"],
    ["", ""],
    ["Broadcast Notifications:", ""],
    ["Public Race Announcement", "Sent to ALL users (batched by 500)"],
]

for row_idx, (label, value) in enumerate(summary_data, 2):
    summary_ws.cell(row=row_idx, column=1).value = label
    summary_ws.cell(row=row_idx, column=2).value = value

    if label and not value:  # Section headers
        summary_ws.cell(row=row_idx, column=1).font = Font(bold=True, size=11)
        summary_ws.merge_cells(f'A{row_idx}:B{row_idx}')

    if label and value and ":" in label:  # Data rows
        summary_ws.cell(row=row_idx, column=1).font = Font(bold=True)

summary_ws.column_dimensions['A'].width = 30
summary_ws.column_dimensions['B'].width = 50

# Save workbook
filename = "Notifications_Documentation.xlsx"
wb.save(filename)

print(f"✅ Excel file created successfully: {filename}")
print(f"📊 Total notifications documented: {len(notifications)}")
print(f"   - Race: {len([n for n in notifications if n['category'] == 'RACE'])}")
print(f"   - Social: {len([n for n in notifications if n['category'] == 'SOCIAL'])}")
print(f"   - Chat: {len([n for n in notifications if n['category'] == 'CHAT'])}")