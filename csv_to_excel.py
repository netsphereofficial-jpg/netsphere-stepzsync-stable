#!/usr/bin/env python3
"""
Convert CSV files to a single Excel workbook with proper formatting
"""

import csv
import os
import sys

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

try:
    import xlsxwriter
    HAS_XLSXWRITER = True
except ImportError:
    HAS_XLSXWRITER = False


def read_csv_file(filepath):
    """Read a CSV file and return rows"""
    with open(filepath, 'r', encoding='utf-8-sig') as f:
        reader = csv.reader(f)
        return list(reader)


def create_excel_with_xlsxwriter(csv_dir, output_file):
    """Create Excel file using xlsxwriter"""
    workbook = xlsxwriter.Workbook(output_file)

    # Define formats
    header_format = workbook.add_format({
        'bold': True,
        'font_color': 'white',
        'bg_color': '#4472C4',
        'border': 1,
        'align': 'center',
        'valign': 'vcenter',
    })

    cell_format = workbook.add_format({
        'border': 1,
        'align': 'left',
        'valign': 'top',
        'text_wrap': True,
    })

    title_format = workbook.add_format({
        'bold': True,
        'font_size': 16,
        'font_color': 'white',
        'bg_color': '#2E75B6',
    })

    section_format = workbook.add_format({
        'bold': True,
        'bg_color': '#D9E1F2',
    })

    # Get all CSV files
    csv_files = sorted([f for f in os.listdir(csv_dir) if f.endswith('.csv')])

    for csv_file in csv_files:
        filepath = os.path.join(csv_dir, csv_file)
        rows = read_csv_file(filepath)

        # Create sheet name from filename
        sheet_name = csv_file.replace('.csv', '').replace('_', ' ')
        # Remove number prefix (e.g., "02 ")
        if sheet_name[0:2].replace(' ', '').isdigit():
            sheet_name = sheet_name[3:]

        # Excel sheet name limit is 31 characters
        sheet_name = sheet_name[:31]

        worksheet = workbook.add_worksheet(sheet_name)

        # Write data
        if 'SUMMARY' in csv_file.upper():
            # Summary sheet - special formatting
            for row_num, row_data in enumerate(rows):
                if row_num == 0:  # Title
                    worksheet.write(row_num, 0, row_data[0], title_format)
                    if len(row_data) > 1:
                        worksheet.write(row_num, 1, row_data[1])
                elif row_data and row_data[0] and row_data[0].endswith(':'):
                    # Section headers
                    worksheet.write(row_num, 0, row_data[0], section_format)
                    if len(row_data) > 1:
                        worksheet.write(row_num, 1, row_data[1], section_format)
                else:
                    for col_num, cell_value in enumerate(row_data):
                        worksheet.write(row_num, col_num, cell_value)

            worksheet.set_column('A:A', 40)
            worksheet.set_column('B:B', 20)
            worksheet.set_column('C:C', 15)
            worksheet.set_column('D:D', 12)

        else:
            # Data sheets
            for row_num, row_data in enumerate(rows):
                if row_num == 0:  # Header row
                    for col_num, cell_value in enumerate(row_data):
                        worksheet.write(row_num, col_num, cell_value, header_format)
                else:
                    for col_num, cell_value in enumerate(row_data):
                        worksheet.write(row_num, col_num, cell_value, cell_format)

            # Set column widths
            worksheet.set_column('A:A', 60)  # English Text
            worksheet.set_column('B:B', 40)  # Screen/Context
            worksheet.set_column('C:C', 50)  # Notes

            # Freeze header row
            worksheet.freeze_panes(1, 0)

            # Add auto-filter
            if len(rows) > 0:
                worksheet.autofilter(0, 0, len(rows) - 1, len(rows[0]) - 1)

    workbook.close()
    return True


def create_excel_with_openpyxl(csv_dir, output_file):
    """Create Excel file using openpyxl"""
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet

    # Define styles
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')

    title_fill = PatternFill(start_color='2E75B6', end_color='2E75B6', fill_type='solid')
    title_font = Font(name='Calibri', size=16, bold=True, color='FFFFFF')

    section_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
    section_font = Font(name='Calibri', size=11, bold=True)

    cell_alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    center_alignment = Alignment(horizontal='center', vertical='center')

    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Get all CSV files
    csv_files = sorted([f for f in os.listdir(csv_dir) if f.endswith('.csv')])

    for csv_file in csv_files:
        filepath = os.path.join(csv_dir, csv_file)
        rows = read_csv_file(filepath)

        # Create sheet name
        sheet_name = csv_file.replace('.csv', '').replace('_', ' ')
        if sheet_name[0:2].replace(' ', '').isdigit():
            sheet_name = sheet_name[3:]
        sheet_name = sheet_name[:31]

        ws = wb.create_sheet(sheet_name)

        # Write data
        for row_num, row_data in enumerate(rows, start=1):
            for col_num, cell_value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_num, column=col_num, value=cell_value)

                # Apply formatting
                if row_num == 1 and 'SUMMARY' not in csv_file.upper():
                    # Header row for data sheets
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = center_alignment
                    cell.border = border
                elif row_num == 1 and 'SUMMARY' in csv_file.upper():
                    # Title for summary
                    cell.fill = title_fill
                    cell.font = title_font
                elif cell_value and isinstance(cell_value, str) and cell_value.endswith(':'):
                    # Section headers
                    cell.fill = section_fill
                    cell.font = section_font
                else:
                    # Regular cells
                    if row_num > 1 and 'SUMMARY' not in csv_file.upper():
                        cell.alignment = cell_alignment
                        cell.border = border

        # Set column widths
        if 'SUMMARY' in csv_file.upper():
            ws.column_dimensions['A'].width = 40
            ws.column_dimensions['B'].width = 20
            ws.column_dimensions['C'].width = 15
            ws.column_dimensions['D'].width = 12
        else:
            ws.column_dimensions['A'].width = 60
            ws.column_dimensions['B'].width = 40
            ws.column_dimensions['C'].width = 50

            # Freeze and filter
            ws.freeze_panes = 'A2'
            if len(rows) > 0:
                ws.auto_filter.ref = ws.dimensions

    wb.save(output_file)
    return True


def main():
    csv_dir = 'translation_sheets'
    output_file = 'StepzSync_Translation_Master.xlsx'

    if not os.path.exists(csv_dir):
        print(f'❌ Error: Directory {csv_dir} not found')
        print('   Please run generate_excel_simple.py first')
        return 1

    print('📊 Converting CSV files to Excel workbook...\n')

    # Try xlsxwriter first, then openpyxl
    if HAS_XLSXWRITER:
        print('   Using xlsxwriter library...')
        try:
            create_excel_with_xlsxwriter(csv_dir, output_file)
            print(f'\n✅ Excel file created: {output_file}')
            return 0
        except Exception as e:
            print(f'❌ Error with xlsxwriter: {e}')

    if HAS_OPENPYXL:
        print('   Using openpyxl library...')
        try:
            create_excel_with_openpyxl(csv_dir, output_file)
            print(f'\n✅ Excel file created: {output_file}')
            return 0
        except Exception as e:
            print(f'❌ Error with openpyxl: {e}')

    # If neither library is available
    print('❌ Error: Neither xlsxwriter nor openpyxl is installed')
    print('   CSV files are available in translation_sheets/ directory')
    print('   You can manually import them into Excel or Google Sheets')
    print('\n   To install dependencies, run:')
    print('   pip3 install xlsxwriter --user')
    return 1


if __name__ == '__main__':
    sys.exit(main())
