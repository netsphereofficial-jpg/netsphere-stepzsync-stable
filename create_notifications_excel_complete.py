#!/usr/bin/env python3
"""
COMPLETE Generate Excel spreadsheet documenting ALL notification types in the app.
This includes race, friend, chat notifications AND ACTIVE IMPLEMENTATIONS from index.js
"""

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

# Create workbook
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "All Notifications"

# Define styles
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=11)
active_fill = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")  # Green for active
pending_fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")  # Yellow for pending
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Set column headers
headers = [
    "No.",
    "Status",
    "Category",
    "Notification Type",
    "Title",
    "Body/Message",
    "Icon",
    "Trigger Event",
    "Trigger Path",
    "Recipients",
    "Additional Data Fields",
    "Notes"
]

# Write headers
for col_num, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_num)
    cell.value = header
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = thin_border

# Set column widths
column_widths = [5, 10, 12, 25, 25, 50, 6, 45, 40, 30, 35, 45]
for i, width in enumerate(column_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = width

# Notification data - COMPLETE LIST
notifications = [
    # ========== RACE NOTIFICATIONS (ACTIVE) ==========
    {
        "status": "ACTIVE",
        "category": "RACE",
        "type": "Race Invitation",
        "title": "Race Invitation 🏃‍♂️",
        "body": "{inviterName} invited you to join \"{raceTitle}\"",
        "icon": "🏃‍♂️",
        "trigger": "Document created in race_invites collection with type='received' and isJoinRequest=false",
        "trigger_path": "functions/notifications/triggers/raceTriggers.js:40-114 (onRaceInviteCreated)",
        "recipients": "Invited user (toUserId)",
        "data_fields": "type: InviteRace, category: Race, raceId, raceName, inviterUserId, inviterName, startTime (optional), distance (optional), location (optional)",
        "notes": "✅ ACTIVE - Triggered when race organizer invites someone to join a race. Only processes 'received' type invites to avoid duplicates."
    },
    {
        "status": "ACTIVE",
        "category": "RACE",
        "type": "New Join Request",
        "title": "New Join Request 🙋‍♂️",
        "body": "{requesterName} wants to join \"{raceTitle}\"",
        "icon": "🙋‍♂️",
        "trigger": "Document created in race_invites collection with type='received' and isJoinRequest=true",
        "trigger_path": "functions/notifications/triggers/raceTriggers.js:40-114 (onRaceInviteCreated)",
        "recipients": "Race organizer (toUserId)",
        "data_fields": "type: NewJoinRequest, category: Race, raceId, raceName, requesterUserId, requesterName, requestedAt",
        "notes": "✅ ACTIVE - Sent when user requests to join a race. Organizer receives this notification."
    },
    {
        "status": "ACTIVE",
        "category": "RACE",
        "type": "Race Started",
        "title": "Race Started! 🚀",
        "body": "\"{raceTitle}\" has begun! Good luck!",
        "icon": "🚀",
        "trigger": "Race document updated with statusId changed to 3 (ACTIVE) - triggered by scheduled function or manual start",
        "trigger_path": "functions/notifications/triggers/raceTriggers.js:122-218 (onRaceStatusChanged)",
        "recipients": "All race participants",
        "data_fields": "type: RaceBegin, category: Race, raceId, raceName, participantCount (optional), startedAt",
        "notes": "✅ ACTIVE - Sent to all participants when race begins. Auto-triggered by autoStartScheduledRaces function or manual start."
    },
    {
        "status": "ACTIVE",
        "category": "RACE",
        "type": "Race Completed (Winner - 1st)",
        "title": "Congratulations! 🥇",
        "body": "You won \"{raceTitle}\"! Amazing performance!",
        "icon": "🏆",
        "trigger": "Race document updated with statusId changed to 4 (COMPLETED)",
        "trigger_path": "functions/notifications/triggers/raceTriggers.js:122-218 (onRaceStatusChanged)",
        "recipients": "Participant who finished 1st",
        "data_fields": "type: RaceWon, category: Achievement, raceId, raceName, rank: 1, xpEarned (optional), distanceCovered (optional), avgSpeed (optional), completedAt",
        "notes": "✅ ACTIVE - Special winner notification for 1st place finisher."
    },
    {
        "status": "ACTIVE",
        "category": "RACE",
        "type": "Race Completed (2nd Place)",
        "title": "Great Job! 🥈",
        "body": "You finished 2nd in \"{raceTitle}\"! Well done!",
        "icon": "🥈",
        "trigger": "Race document updated with statusId changed to 4 (COMPLETED)",
        "trigger_path": "functions/notifications/triggers/raceTriggers.js:122-218 (onRaceStatusChanged)",
        "recipients": "Participant who finished 2nd",
        "data_fields": "type: RaceCompleted, category: Achievement, raceId, raceName, rank: 2, xpEarned (optional), distanceCovered (optional), avgSpeed (optional), completedAt",
        "notes": "✅ ACTIVE - Sent to 2nd place finisher with silver medal emoji."
    },
    {
        "status": "ACTIVE",
        "category": "RACE",
        "type": "Race Completed (3rd Place)",
        "title": "Excellent! 🥉",
        "body": "You finished 3rd in \"{raceTitle}\"! Great effort!",
        "icon": "🥉",
        "trigger": "Race document updated with statusId changed to 4 (COMPLETED)",
        "trigger_path": "functions/notifications/triggers/raceTriggers.js:122-218 (onRaceStatusChanged)",
        "recipients": "Participant who finished 3rd",
        "data_fields": "type: RaceCompleted, category: Achievement, raceId, raceName, rank: 3, xpEarned (optional), distanceCovered (optional), avgSpeed (optional), completedAt",
        "notes": "✅ ACTIVE - Sent to 3rd place finisher with bronze medal emoji."
    },
    {
        "status": "ACTIVE",
        "category": "RACE",
        "type": "Race Completed (Other)",
        "title": "Race Completed! 🏃‍♂️",
        "body": "You finished \"{raceTitle}\" in {rank} place!",
        "icon": "🏃‍♂️",
        "trigger": "Race document updated with statusId changed to 4 (COMPLETED)",
        "trigger_path": "functions/notifications/triggers/raceTriggers.js:122-218 (onRaceStatusChanged)",
        "recipients": "Participants who finished 4th or lower",
        "data_fields": "type: RaceCompleted, category: Race, raceId, raceName, rank, xpEarned (optional), distanceCovered (optional), avgSpeed (optional), completedAt",
        "notes": "✅ ACTIVE - Sent to all other finishers with their rank (4th, 5th, etc.)."
    },
    {
        "status": "ACTIVE",
        "category": "RACE",
        "type": "First Finisher",
        "title": "🏁 First to Finish!",
        "body": "Amazing! You're the first to complete \"{raceTitle}\"!",
        "icon": "🏁",
        "trigger": "Race document updated with statusId changed to 6 (ENDING) - first participant crosses finish",
        "trigger_path": "functions/notifications/triggers/raceTriggers.js:154-200 (onRaceStatusChanged)",
        "recipients": "First finisher (firstFinisherUserId)",
        "data_fields": "type: RaceFirstFinisher, category: Achievement, raceId, raceName, finishedAt",
        "notes": "✅ ACTIVE - Sent when first participant completes the race, triggering the deadline countdown."
    },
    {
        "status": "ACTIVE",
        "category": "RACE",
        "type": "Deadline Alert",
        "title": "⏰ Deadline Approaching!",
        "body": "{firstFinisherName} finished first! You have {deadlineMinutes} minutes to complete the race!",
        "icon": "⏰",
        "trigger": "Race document updated with statusId changed to 6 (ENDING) - deadline countdown starts",
        "trigger_path": "functions/notifications/triggers/raceTriggers.js:154-200 (onRaceStatusChanged)",
        "recipients": "All active participants who haven't finished yet",
        "data_fields": "type: RaceDeadlineAlert, category: Race, raceId, raceName, firstFinisherName, deadlineMinutes, deadline (ISO timestamp), timestamp",
        "notes": "✅ ACTIVE - Sent to remaining active participants when first person finishes, creates urgency to complete."
    },
    {
        "status": "ACTIVE",
        "category": "RACE",
        "type": "Race Cancelled",
        "title": "❌ Race Cancelled",
        "body": "The race \"{raceTitle}\" has been cancelled. Reason: {reason}",
        "icon": "❌",
        "trigger": "Race document updated with statusId changed to 7 (CANCELLED)",
        "trigger_path": "functions/notifications/triggers/raceTriggers.js:202-211 (onRaceStatusChanged)",
        "recipients": "All race participants",
        "data_fields": "type: RaceCancelled, category: Race, raceId, raceName, cancellationReason, cancelledAt",
        "notes": "✅ ACTIVE - Sent when race organizer cancels the race. Includes cancellation reason if provided."
    },
    {
        "status": "ACTIVE",
        "category": "RACE",
        "type": "Join Request Accepted",
        "title": "Join Request Accepted ✅",
        "body": "{organizerName} accepted your request to join \"{raceTitle}\"",
        "icon": "✅",
        "trigger": "race_invites document updated with status='accepted' and isJoinRequest=true",
        "trigger_path": "functions/notifications/triggers/raceTriggers.js:226-300 (onRaceInviteAccepted)",
        "recipients": "User who requested to join (toUserId)",
        "data_fields": "type: JoinRequestAccepted, category: Race, raceId, raceName, organizerUserId, organizerName, acceptedAt",
        "notes": "✅ ACTIVE - Sent when organizer approves a join request."
    },
    {
        "status": "ACTIVE",
        "category": "RACE",
        "type": "Invite Accepted",
        "title": "Race Invite Accepted 🎉",
        "body": "{accepterName} accepted your invite to \"{raceTitle}\"",
        "icon": "🎉",
        "trigger": "race_invites document updated with status='accepted' and isJoinRequest=false",
        "trigger_path": "functions/notifications/triggers/raceTriggers.js:226-300 (onRaceInviteAccepted)",
        "recipients": "Race organizer who sent invite (fromUserId)",
        "data_fields": "type: InviteAccepted, category: Race, raceId, raceName, accepterUserId, accepterName, acceptedAt",
        "notes": "✅ ACTIVE - Sent to organizer when invited user accepts race invitation."
    },
    {
        "status": "ACTIVE",
        "category": "RACE",
        "type": "Join Request Declined",
        "title": "Join Request Declined",
        "body": "{organizerName} declined your request to join \"{raceTitle}\"",
        "icon": "❌",
        "trigger": "race_invites document updated with status='declined' and isJoinRequest=true",
        "trigger_path": "functions/notifications/triggers/raceTriggers.js:308-382 (onRaceInviteDeclined)",
        "recipients": "User who requested to join (toUserId)",
        "data_fields": "type: JoinRequestDeclined, category: Race, raceId, raceName, organizerUserId, organizerName, declinedAt",
        "notes": "✅ ACTIVE - Sent when organizer rejects a join request."
    },
    {
        "status": "ACTIVE",
        "category": "RACE",
        "type": "Invite Declined",
        "title": "Race Invite Declined",
        "body": "{declinerName} declined your invite to \"{raceTitle}\"",
        "icon": "😔",
        "trigger": "race_invites document updated with status='declined' and isJoinRequest=false",
        "trigger_path": "functions/notifications/triggers/raceTriggers.js:308-382 (onRaceInviteDeclined)",
        "recipients": "Race organizer who sent invite (fromUserId)",
        "data_fields": "type: InviteDeclined, category: Race, raceId, raceName, declinerUserId, declinerName, declinedAt",
        "notes": "✅ ACTIVE - Sent to organizer when invited user declines race invitation."
    },
    {
        "status": "ACTIVE",
        "category": "RACE",
        "type": "Race Created (Confirmation)",
        "title": "Race Created Successfully! 🎉",
        "body": "Your {raceType} \"{raceTitle}\" is ready! Distance: {distance}km.",
        "icon": "🎉",
        "trigger": "New document created in races collection",
        "trigger_path": "functions/notifications/triggers/raceTriggers.js:399-443 (onRaceCreated)",
        "recipients": "Race creator (createdBy)",
        "data_fields": "type: RaceCreated, category: Race, raceId, raceName, raceType, distance, scheduledTime (optional), participantCount (optional), createdAt",
        "notes": "✅ ACTIVE - Confirmation sent to race creator immediately after race creation."
    },
    {
        "status": "ACTIVE",
        "category": "RACE",
        "type": "Public Race Announcement",
        "title": "New Public Race Available! 🏁",
        "body": "{organizerName} created \"{raceTitle}\" - {distance}km. Join now!",
        "icon": "🏁",
        "trigger": "New document created in races collection with raceTypeId=3 (Public)",
        "trigger_path": "functions/notifications/triggers/raceTriggers.js:399-443 (onRaceCreated) + senders/raceNotifications.js:602-735 (sendPublicRaceAnnouncement)",
        "recipients": "ALL app users (except creator)",
        "data_fields": "type: PublicRaceAnnouncement, category: Race, raceId, raceName, raceTypeId: 3, distance, organizerName, location (optional), scheduledTime (optional), participantLimit (optional), createdAt",
        "notes": "✅ ACTIVE - BROADCAST notification sent to all users when public race is created. Sent in batches of 500 users via FCM."
    },
    {
        "status": "ACTIVE",
        "category": "RACE",
        "type": "Participant Joined",
        "title": "Someone Joined Your Race! 🎉",
        "body": "{participantName} joined \"{raceTitle}\"",
        "icon": "🎉",
        "trigger": "Document created in races/{raceId}/participants/{userId} subcollection",
        "trigger_path": "functions/index.js:26-93 (onParticipantJoined) + senders/raceNotifications.js:759-792",
        "recipients": "Race organizer/creator (excludes if participant is organizer)",
        "data_fields": "type: RaceParticipantJoined, category: Race, raceId, raceName, participantId, participantName, joinedAt",
        "notes": "✅ ACTIVE - Triggered when someone joins a race. Notifies race organizer. Also increments participantCount."
    },
    {
        "status": "ACTIVE",
        "category": "RACE",
        "type": "Overtaking (Overtaker)",
        "title": "Great Overtake! 🚀",
        "body": "Awesome! You overtook {overtakenName} and moved to rank #{newRank}!",
        "icon": "🚀",
        "trigger": "Participant document updated with improved rank (lower rank number)",
        "trigger_path": "functions/index.js:147-336 (onParticipantUpdated lines 182-221) + senders/raceNotifications.js:798-903",
        "recipients": "Participant who overtook",
        "data_fields": "type: RaceOvertaking, category: Achievement, raceId, raceName, newRank, oldRank, overtakenUser, timestamp",
        "notes": "✅ ACTIVE - Triggered when user improves rank during race. Sends positive reinforcement notification."
    },
    {
        "status": "ACTIVE",
        "category": "RACE",
        "type": "Overtaking (Overtaken)",
        "title": "You Were Overtaken! ⚡",
        "body": "{overtakerName} just overtook you! Speed up to reclaim your position!",
        "icon": "⚡",
        "trigger": "Participant document updated (detected when another participant overtakes)",
        "trigger_path": "functions/index.js:147-336 (onParticipantUpdated lines 182-221) + senders/raceNotifications.js:798-903",
        "recipients": "Participant who was overtaken",
        "data_fields": "type: RaceOvertaken, category: Race, raceId, raceName, overtakerName, yourNewRank, timestamp",
        "notes": "✅ ACTIVE - Triggered when user is overtaken. Creates competitive pressure to speed up."
    },
    {
        "status": "ACTIVE",
        "category": "RACE",
        "type": "Overtaking (General Alert)",
        "title": "Overtaking Alert! 🏃‍♂️",
        "body": "{overtakerName} overtook {overtakenName} and moved to rank #{newRank}!",
        "icon": "🏃‍♂️",
        "trigger": "Participant document updated (rank change detected)",
        "trigger_path": "functions/index.js:147-336 (onParticipantUpdated lines 182-221) + senders/raceNotifications.js:798-903",
        "recipients": "All other race participants (except overtaker and overtaken, excludes winners)",
        "data_fields": "type: RaceOvertakingGeneral, category: Race, raceId, raceName, overtakingUser, overtakenUser, newRank, timestamp",
        "notes": "✅ ACTIVE - Keeps all participants informed of race dynamics. Creates competitive atmosphere."
    },
    {
        "status": "ACTIVE",
        "category": "RACE",
        "type": "Leader Change",
        "title": "New Leader! 👑",
        "body": "{newLeaderName} took the lead in \"{raceTitle}\"!",
        "icon": "👑",
        "trigger": "Participant document updated with rank changed to 1 (from non-1)",
        "trigger_path": "functions/index.js:147-336 (onParticipantUpdated lines 224-263) + senders/raceNotifications.js:909-959",
        "recipients": "All race participants (except new leader, excludes winners)",
        "data_fields": "type: RaceLeaderChange, category: Race, raceId, raceName, newLeaderUserId, newLeaderName, timestamp",
        "notes": "✅ ACTIVE - Triggered when someone takes 1st place. Also updates topParticipant in race document."
    },
    {
        "status": "ACTIVE",
        "category": "RACE",
        "type": "Personal Milestone (25%)",
        "title": "Milestone Reached! 🎯",
        "body": "Great job! You've completed 25% of \"{raceTitle}\"!",
        "icon": "🎯",
        "trigger": "Participant document updated with distance crossing 25% threshold",
        "trigger_path": "functions/index.js:147-336 (onParticipantUpdated lines 265-321) + senders/raceNotifications.js:1121-1156",
        "recipients": "Participant who reached milestone",
        "data_fields": "type: RaceMilestonePersonal, category: Achievement, raceId, raceName, milestone: 25, achievedAt",
        "notes": "✅ ACTIVE - Personal achievement notification for 25% completion. Prevents duplicate notifications using reachedMilestones array."
    },
    {
        "status": "ACTIVE",
        "category": "RACE",
        "type": "Personal Milestone (50%)",
        "title": "Milestone Reached! ⚡",
        "body": "Great job! You've completed 50% of \"{raceTitle}\"!",
        "icon": "⚡",
        "trigger": "Participant document updated with distance crossing 50% threshold",
        "trigger_path": "functions/index.js:147-336 (onParticipantUpdated lines 265-321) + senders/raceNotifications.js:1121-1156",
        "recipients": "Participant who reached milestone",
        "data_fields": "type: RaceMilestonePersonal, category: Achievement, raceId, raceName, milestone: 50, achievedAt",
        "notes": "✅ ACTIVE - Personal achievement notification for 50% completion (halfway point)."
    },
    {
        "status": "ACTIVE",
        "category": "RACE",
        "type": "Personal Milestone (75%)",
        "title": "Milestone Reached! 🔥",
        "body": "Great job! You've completed 75% of \"{raceTitle}\"!",
        "icon": "🔥",
        "trigger": "Participant document updated with distance crossing 75% threshold",
        "trigger_path": "functions/index.js:147-336 (onParticipantUpdated lines 265-321) + senders/raceNotifications.js:1121-1156",
        "recipients": "Participant who reached milestone",
        "data_fields": "type: RaceMilestonePersonal, category: Achievement, raceId, raceName, milestone: 75, achievedAt",
        "notes": "✅ ACTIVE - Personal achievement notification for 75% completion (almost done!)."
    },
    {
        "status": "ACTIVE",
        "category": "RACE",
        "type": "Milestone Alert (25%)",
        "title": "{userName} Hit 25%! 🎯",
        "body": "{userName} reached 25% of \"{raceTitle}\". Keep pushing!",
        "icon": "🎯",
        "trigger": "Participant document updated (someone else reached 25%)",
        "trigger_path": "functions/index.js:147-336 (onParticipantUpdated lines 265-321) + senders/raceNotifications.js:1162-1226",
        "recipients": "All other race participants (except achiever, excludes winners)",
        "data_fields": "type: RaceMilestoneAlert, category: Race, raceId, raceName, achieverName, achieverUserId, milestone: 25, timestamp",
        "notes": "✅ ACTIVE - Informs other participants when someone hits 25% milestone."
    },
    {
        "status": "ACTIVE",
        "category": "RACE",
        "type": "Milestone Alert (50%)",
        "title": "{userName} Hit 50%! ⚡",
        "body": "{userName} reached 50% of \"{raceTitle}\". Keep pushing!",
        "icon": "⚡",
        "trigger": "Participant document updated (someone else reached 50%)",
        "trigger_path": "functions/index.js:147-336 (onParticipantUpdated lines 265-321) + senders/raceNotifications.js:1162-1226",
        "recipients": "All other race participants (except achiever, excludes winners)",
        "data_fields": "type: RaceMilestoneAlert, category: Race, raceId, raceName, achieverName, achieverUserId, milestone: 50, timestamp",
        "notes": "✅ ACTIVE - Informs other participants when someone hits 50% milestone (halfway point)."
    },
    {
        "status": "ACTIVE",
        "category": "RACE",
        "type": "Milestone Alert (75%)",
        "title": "{userName} Hit 75%! 🔥",
        "body": "{userName} reached 75% of \"{raceTitle}\". Keep pushing!",
        "icon": "🔥",
        "trigger": "Participant document updated (someone else reached 75%)",
        "trigger_path": "functions/index.js:147-336 (onParticipantUpdated lines 265-321) + senders/raceNotifications.js:1162-1226",
        "recipients": "All other race participants (except achiever, excludes winners)",
        "data_fields": "type: RaceMilestoneAlert, category: Race, raceId, raceName, achieverName, achieverUserId, milestone: 75, timestamp",
        "notes": "✅ ACTIVE - Informs other participants when someone hits 75% milestone (almost done!)."
    },
    {
        "status": "ACTIVE",
        "category": "RACE",
        "type": "Countdown Timer (5 minutes)",
        "title": "⏰ 5 Minutes Left!",
        "body": "Time is running out in \"{raceTitle}\"! Sprint to the finish!",
        "icon": "⏰",
        "trigger": "Scheduled function runs every 1 minute, checks races in ENDING status (statusId=6) with deadline 4-5 minutes away",
        "trigger_path": "functions/scheduled/raceCountdownChecker.js:28-115 (checkRaceCountdowns) + senders/raceNotifications.js:1279-1333",
        "recipients": "All active participants who haven't finished",
        "data_fields": "type: RaceCountdownTimer, category: Race, raceId, raceName, minutesLeft: 5, timestamp",
        "notes": "✅ ACTIVE - Scheduled function runs every minute. Sends when 5 minutes remaining. Prevents duplicate with countdownNotificationSent flag."
    },

    # ========== RACE NOTIFICATIONS (PENDING - Not Implemented) ==========
    {
        "status": "PENDING",
        "category": "RACE",
        "type": "Proximity Alert",
        "title": "🔥 Opponent Approaching!",
        "body": "{chaserName} is only {distanceGap}m behind you! Speed up!",
        "icon": "🔥",
        "trigger": "NOT IMPLEMENTED - Would trigger when opponent gets within 20m",
        "trigger_path": "functions/notifications/senders/raceNotifications.js:1232-1273 (sendProximityAlertNotification) - Function exists but no trigger",
        "recipients": "Participant being chased",
        "data_fields": "type: RaceProximityAlert, category: Race, raceId, raceName, chaserName, chaserUserId, distanceGap (meters), timestamp",
        "notes": "⚠️ PENDING - Function exists but trigger not implemented. Would alert when opponent closes within 20m gap."
    },
    {
        "status": "PENDING",
        "category": "RACE",
        "type": "Race Reminder (15 min)",
        "title": "Race Starting Soon! ⏰",
        "body": "\"{raceTitle}\" starts in 15 minutes. Get ready!",
        "icon": "⏰",
        "trigger": "NOT IMPLEMENTED - Would need scheduled function to check upcoming races 15 minutes before start",
        "trigger_path": "functions/notifications/senders/raceNotifications.js:214-268 (sendRaceReminder) - Function exists but no trigger",
        "recipients": "All race participants",
        "data_fields": "type: RaceReminder, category: Race, raceId, raceName, reminderType: 15min, startTime (optional), reminderSentAt",
        "notes": "⚠️ PENDING - Function exists but trigger not implemented. Would be triggered 15 minutes before race start."
    },
    {
        "status": "PENDING",
        "category": "RACE",
        "type": "Race Reminder (1 hour)",
        "title": "Race Reminder 🕐",
        "body": "\"{raceTitle}\" starts in 1 hour. Don't forget!",
        "icon": "⏰",
        "trigger": "NOT IMPLEMENTED - Would need scheduled function to check upcoming races 1 hour before start",
        "trigger_path": "functions/notifications/senders/raceNotifications.js:214-268 (sendRaceReminder) - Function exists but no trigger",
        "recipients": "All race participants",
        "data_fields": "type: RaceReminder, category: Race, raceId, raceName, reminderType: 1hour, startTime (optional), reminderSentAt",
        "notes": "⚠️ PENDING - Function exists but trigger not implemented. Would be triggered 1 hour before race start."
    },
    {
        "status": "PENDING",
        "category": "RACE",
        "type": "Race Reminder (1 day)",
        "title": "Race Tomorrow 📅",
        "body": "\"{raceTitle}\" is scheduled for tomorrow.",
        "icon": "⏰",
        "trigger": "NOT IMPLEMENTED - Would need scheduled function to check upcoming races 1 day before start",
        "trigger_path": "functions/notifications/senders/raceNotifications.js:214-268 (sendRaceReminder) - Function exists but no trigger",
        "recipients": "All race participants",
        "data_fields": "type: RaceReminder, category: Race, raceId, raceName, reminderType: 1day, startTime (optional), reminderSentAt",
        "notes": "⚠️ PENDING - Function exists but trigger not implemented. Would be triggered 1 day before race start."
    },

    # ========== FRIEND/SOCIAL NOTIFICATIONS (ACTIVE) ==========
    {
        "status": "ACTIVE",
        "category": "SOCIAL",
        "type": "Friend Request",
        "title": "New Friend Request 👥",
        "body": "{senderName} wants to be your friend!",
        "icon": "👥",
        "trigger": "Document created in friend_requests collection",
        "trigger_path": "functions/notifications/triggers/friendTriggers.js:30-73 (onFriendRequestCreated)",
        "recipients": "Receiver of friend request",
        "data_fields": "type: FriendRequest, category: Social, userId (sender), userName, thumbnail (profile pic, optional), mutualFriends (optional), requestSentAt",
        "notes": "✅ ACTIVE - Sent when user sends a friend request to another user."
    },
    {
        "status": "ACTIVE",
        "category": "SOCIAL",
        "type": "Friend Request Accepted",
        "title": "Friend Request Accepted! 🎉",
        "body": "{friendName} accepted your friend request!",
        "icon": "🎉",
        "trigger": "friend_requests document updated with status='accepted'",
        "trigger_path": "functions/notifications/triggers/friendTriggers.js:81-126 (onFriendRequestAccepted)",
        "recipients": "Original friend request sender",
        "data_fields": "type: FriendAccepted, category: Social, userId (accepter), userName, thumbnail (profile pic, optional), acceptedAt",
        "notes": "✅ ACTIVE - Sent to requester when receiver accepts the friend request."
    },
    {
        "status": "ACTIVE",
        "category": "SOCIAL",
        "type": "Friend Request Declined",
        "title": "Friend Request Declined 😔",
        "body": "{friendName} declined your friend request.",
        "icon": "😔",
        "trigger": "friend_requests document updated with status='declined'",
        "trigger_path": "functions/notifications/triggers/friendTriggers.js:134-178 (onFriendRequestDeclined)",
        "recipients": "Original friend request sender",
        "data_fields": "type: FriendDeclined, category: Social, userId (decliner), userName, declinedAt",
        "notes": "✅ ACTIVE - Sent to requester when receiver declines the friend request."
    },
    {
        "status": "ACTIVE",
        "category": "SOCIAL",
        "type": "Friend Removed",
        "title": "Friendship Ended 💔",
        "body": "{removerName} removed you from their friends list.",
        "icon": "💔",
        "trigger": "Document deleted from friends collection",
        "trigger_path": "functions/notifications/triggers/friendTriggers.js:190-228 (onFriendRemoved)",
        "recipients": "Friend who was removed",
        "data_fields": "type: FriendRemoved, category: Social, userId (remover), userName, removedAt",
        "notes": "✅ ACTIVE - Sent when user removes another user from friends list. Only sent to friendId user to avoid duplicate notifications."
    },

    # ========== CHAT NOTIFICATIONS (ACTIVE) ==========
    {
        "status": "ACTIVE",
        "category": "CHAT",
        "type": "Direct Chat Message",
        "title": "New Message from {senderName} 💬",
        "body": "{messageText}",
        "icon": "💬",
        "trigger": "Document created in chat_messages collection",
        "trigger_path": "functions/notifications/triggers/chatTriggers.js:26-53 (onChatMessageCreated)",
        "recipients": "Message receiver (receiverId)",
        "data_fields": "type: ChatMessage, category: Chat, senderName, chatRoomId, clickAction: OPEN_CHAT",
        "notes": "✅ ACTIVE - Sent for 1-on-1 direct messages. Message truncated to 100 chars in notification."
    },
    {
        "status": "ACTIVE",
        "category": "CHAT",
        "type": "Race Chat Message",
        "title": "{raceTitle} 🏃",
        "body": "{senderName}: {messageText}",
        "icon": "🏃",
        "trigger": "Document created in race_chat_messages collection",
        "trigger_path": "functions/notifications/triggers/chatTriggers.js:61-119 (onRaceChatMessageCreated)",
        "recipients": "All race chat participants (except sender)",
        "data_fields": "type: RaceChatMessage, category: RaceChat, senderName, raceTitle, raceId, raceChatId, clickAction: OPEN_RACE_CHAT",
        "notes": "✅ ACTIVE - Sent to all race chat participants when someone sends a message. Message truncated to 80 chars. Notifications sent in parallel to all participants."
    },
]

# Write data rows
row_num = 2
for idx, notif in enumerate(notifications, 1):
    # Write data
    ws.cell(row=row_num, column=1).value = idx
    ws.cell(row=row_num, column=2).value = notif["status"]
    ws.cell(row=row_num, column=3).value = notif["category"]
    ws.cell(row=row_num, column=4).value = notif["type"]
    ws.cell(row=row_num, column=5).value = notif["title"]
    ws.cell(row=row_num, column=6).value = notif["body"]
    ws.cell(row=row_num, column=7).value = notif["icon"]
    ws.cell(row=row_num, column=8).value = notif["trigger"]
    ws.cell(row=row_num, column=9).value = notif["trigger_path"]
    ws.cell(row=row_num, column=10).value = notif["recipients"]
    ws.cell(row=row_num, column=11).value = notif["data_fields"]
    ws.cell(row=row_num, column=12).value = notif["notes"]

    # Apply styling
    for col in range(1, 13):
        cell = ws.cell(row=row_num, column=col)
        cell.alignment = Alignment(vertical='top', wrap_text=True)
        cell.border = thin_border

        # Apply color coding based on status
        if col == 2:  # Status column
            if notif["status"] == "ACTIVE":
                cell.fill = active_fill
                cell.font = Font(bold=True, color="006100")
            elif notif["status"] == "PENDING":
                cell.fill = pending_fill
                cell.font = Font(bold=True, color="9C6500")

        # Bold category column
        if col == 3:
            cell.font = Font(bold=True)

        # Center number and status columns
        if col in [1, 2]:
            cell.alignment = Alignment(horizontal='center', vertical='top')

    row_num += 1

# Set row heights
ws.row_dimensions[1].height = 30
for row in range(2, row_num):
    ws.row_dimensions[row].height = 65

# Freeze header row and first 3 columns
ws.freeze_panes = 'D2'

# Add summary sheet
summary_ws = wb.create_sheet("Summary & Statistics", 0)

# Summary header
summary_ws.merge_cells('A1:B1')
summary_ws['A1'] = "🔔 Notification System - Complete Documentation"
summary_ws['A1'].font = Font(bold=True, size=14, color="FFFFFF")
summary_ws['A1'].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
summary_ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
summary_ws.row_dimensions[1].height = 25

# Summary data
active_count = len([n for n in notifications if n["status"] == "ACTIVE"])
pending_count = len([n for n in notifications if n["status"] == "PENDING"])
race_active = len([n for n in notifications if n["category"] == "RACE" and n["status"] == "ACTIVE"])
race_pending = len([n for n in notifications if n["category"] == "RACE" and n["status"] == "PENDING"])

summary_data = [
    ["Generated On:", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
    ["Project:", "NetSphere StepZSync"],
    ["", ""],
    ["📊 OVERALL STATISTICS", ""],
    ["Total Notifications:", len(notifications)],
    ["✅ Active (Implemented):", active_count],
    ["⚠️ Pending (Not Triggered):", pending_count],
    ["", ""],
    ["📈 BREAKDOWN BY CATEGORY", ""],
    ["Race Notifications:", f"{race_active} active + {race_pending} pending = {race_active + race_pending} total"],
    ["Social Notifications:", f"{len([n for n in notifications if n['category'] == 'SOCIAL'])} (all active)"],
    ["Chat Notifications:", f"{len([n for n in notifications if n['category'] == 'CHAT'])} (all active)"],
    ["", ""],
    ["🔥 ACTIVE TRIGGERS (Deployed)", ""],
    ["race_invites (onCreate)", "Race invitations & join requests"],
    ["races (onUpdate - statusId)", "Race status changes (started, completed, cancelled, ending)"],
    ["races (onCreate)", "Race creation & public race broadcasts"],
    ["race_invites (onUpdate)", "Invite/join request acceptance/decline"],
    ["races/{id}/participants (onCreate)", "Participant joined notifications"],
    ["races/{id}/participants (onUpdate)", "Overtaking, leader change, milestones (25%, 50%, 75%)"],
    ["friend_requests (onCreate)", "Friend requests"],
    ["friend_requests (onUpdate)", "Friend request responses (accepted/declined)"],
    ["friends (onDelete)", "Friend removal"],
    ["chat_messages (onCreate)", "Direct 1-on-1 chat messages"],
    ["race_chat_messages (onCreate)", "Race group chat messages"],
    ["", ""],
    ["⏰ SCHEDULED FUNCTIONS (Active)", ""],
    ["autoStartScheduledRaces", "Runs every 1 minute - Auto-starts races at scheduled time (statusId 1→3)"],
    ["checkRaceCountdowns", "Runs every 1 minute - Sends 5-minute countdown alerts for ending races"],
    ["", ""],
    ["🚀 KEY FEATURES", ""],
    ["Broadcast Notifications", "Public race announcements sent to ALL users (batched by 500)"],
    ["Real-time Race Updates", "Overtaking, leader changes, milestone alerts during active races"],
    ["Milestone Tracking", "Personal + alert notifications at 25%, 50%, 75% completion"],
    ["Duplicate Prevention", "Uses flags (countdownNotificationSent, reachedMilestones array)"],
    ["Message Truncation", "100 chars (direct chat), 80 chars (race chat)"],
    ["Winner Exclusion", "Completed winners don't receive ongoing race alerts"],
    ["", ""],
    ["⚠️ PENDING IMPLEMENTATIONS", ""],
    ["Proximity Alerts", "Function ready - needs trigger for <20m opponent distance"],
    ["Race Reminders", "Functions ready - need scheduled checks (15min, 1hr, 1day before start)"],
    ["", ""],
    ["📝 IMPORTANT NOTES", ""],
    ["Cloud Functions Deployed:", "functions/index.js exports all triggers"],
    ["FCM Service:", "functions/notifications/core/fcmService.js handles all FCM messaging"],
    ["Notification Channels:", "race_notifications (Android), default sound + badge (iOS)"],
    ["Error Handling:", "Notifications failures don't break main operations (try-catch wrapped)"],
]

for row_idx, (label, value) in enumerate(summary_data, 2):
    summary_ws.cell(row=row_idx, column=1).value = label
    summary_ws.cell(row=row_idx, column=2).value = value

    if label and not value:  # Section headers
        summary_ws.cell(row=row_idx, column=1).font = Font(bold=True, size=12)
        summary_ws.merge_cells(f'A{row_idx}:B{row_idx}')
        if "STATISTICS" in label or "BREAKDOWN" in label:
            summary_ws.cell(row=row_idx, column=1).fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")

    if label and value and ":" in label:  # Data rows
        summary_ws.cell(row=row_idx, column=1).font = Font(bold=True)

summary_ws.column_dimensions['A'].width = 35
summary_ws.column_dimensions['B'].width = 70

# Save workbook
filename = "Notifications_Documentation_COMPLETE.xlsx"
wb.save(filename)

print(f"✅ COMPLETE Excel file created: {filename}")
print(f"📊 Total notifications documented: {len(notifications)}")
print(f"   ✅ ACTIVE: {active_count}")
print(f"   ⚠️ PENDING: {pending_count}")
print(f"\n📈 By Category:")
print(f"   - Race: {race_active} active + {race_pending} pending = {race_active + race_pending} total")
print(f"   - Social: {len([n for n in notifications if n['category'] == 'SOCIAL'])} (all active)")
print(f"   - Chat: {len([n for n in notifications if n['category'] == 'CHAT'])} (all active)")
