#!/usr/bin/env python3
"""
Generate Excel translation workbook from extracted strings
"""

import json
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

def load_strings_data():
    """Load the extracted strings from JSON file"""
    with open('translation_strings.json', 'r', encoding='utf-8') as f:
        data = json.load(f)
    return data

def clean_and_deduplicate_strings(strings_data):
    """Clean strings and remove obvious duplicates while preserving context"""
    seen = {}
    cleaned = []

    for string_obj in strings_data:
        text = string_obj['text']
        screen = string_obj['screenContext']
        category = string_obj['category']

        # Skip very technical strings that slipped through
        if _is_technical_string(text):
            continue

        # Create a key for deduplication
        key = (text.lower().strip(), category)

        if key not in seen:
            seen[key] = string_obj
            cleaned.append(string_obj)
        else:
            # If same text in different screens, combine the contexts
            existing = seen[key]
            if screen not in existing['screenContext']:
                existing['screenContext'] += f'; {screen}'

    return cleaned

def _is_technical_string(text):
    """Filter out technical/non-user-facing strings"""
    technical_keywords = [
        'widget', 'controller', 'service', 'model', 'provider',
        'firebase', 'firestore', 'collection', 'document',
        '.dart', '.json', 'toString', 'override', 'async',
        'await', 'class', 'extends', 'implements', 'void',
        'String', 'int', 'bool', 'double', 'List', 'Map',
        'setState', 'initState', 'dispose', 'build',
    ]

    lower_text = text.lower()

    # Check for technical keywords
    for keyword in technical_keywords:
        if keyword in lower_text:
            return True

    # Check for camelCase (likely variable names)
    if any(c.isupper() for c in text[1:]) and not any(c.isspace() for c in text):
        return True

    # Check for snake_case with multiple underscores
    if text.count('_') > 1:
        return True

    # Very short technical strings
    if len(text) < 3 and not text.isalpha():
        return True

    return False

def organize_by_sheets(strings_data):
    """Organize strings into sheets by category"""
    sheets_data = {}

    category_order = [
        'Authentication',
        'Profile & Settings',
        'Race Management',
        'Active Races',
        'Social Features',
        'Leaderboard & Stats',
        'Home & Navigation',
        'Dialogs & Popups',
        'Subscription/Premium',
        'Errors & Validation',
        'Admin Dashboard',
        'Common/Shared',
    ]

    for category in category_order:
        sheets_data[category] = []

    for string_obj in strings_data:
        category = string_obj['category']
        sheets_data[category].append({
            'English Text': string_obj['text'],
            'Screen/Context': string_obj['screenContext'],
            'Notes': string_obj['notes'],
        })

    return sheets_data

def create_excel_workbook(sheets_data, metadata):
    """Create Excel workbook with all sheets"""

    # Create Excel writer
    output_file = 'StepzSync_Translation_Master.xlsx'

    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:

        # Create Instructions sheet first
        create_instructions_sheet(writer, metadata)

        # Create Summary sheet
        create_summary_sheet(writer, sheets_data, metadata)

        # Create category sheets
        for sheet_name, strings in sheets_data.items():
            if strings:  # Only create sheet if it has data
                df = pd.DataFrame(strings)
                sheet_display_name = sheet_name[:31]  # Excel sheet name limit
                df.to_excel(writer, sheet_name=sheet_display_name, index=False)

    # Now apply formatting
    apply_formatting(output_file, sheets_data)

    print(f'\n✅ Excel file created: {output_file}')
    return output_file

def create_instructions_sheet(writer, metadata):
    """Create instructions sheet for translators"""
    instructions = [
        ['StepzSync Translation Guide', ''],
        ['', ''],
        ['Project Information:', ''],
        ['App Name:', 'StepzSync'],
        ['Type:', 'Fitness Tracking & Racing App'],
        ['Platform:', 'Flutter (iOS & Android)'],
        ['Extracted Date:', metadata['extractedAt'][:10]],
        ['Total Strings:', metadata['totalStrings']],
        ['', ''],
        ['Instructions for Translators:', ''],
        ['', ''],
        ['1. How to Use This Workbook:', ''],
        ['   - Each sheet represents a different section of the app', ''],
        ['   - Column A: English Text - The original text to translate', ''],
        ['   - Column B: Screen/Context - Where this text appears in the app', ''],
        ['   - Column C: Notes - Additional context about the string', ''],
        ['', ''],
        ['2. Translation Guidelines:', ''],
        ['   - Keep the same tone and style as the English text', ''],
        ['   - For button labels, keep translations concise', ''],
        ['   - Pay attention to placeholders (marked in Notes)', ''],
        ['   - Placeholders like {count}, {name}, $variable should NOT be translated', ''],
        ['   - Maintain similar text length for UI elements', ''],
        ['   - Error messages should be clear and helpful', ''],
        ['   - Questions should end with appropriate punctuation', ''],
        ['', ''],
        ['3. Special Considerations:', ''],
        ['   - Health & Fitness terms should use standard translations', ''],
        ['   - Race terminology should be consistent throughout', ''],
        ['   - Premium/Subscription text should be clear about benefits', ''],
        ['   - Permission requests should explain why clearly', ''],
        ['', ''],
        ['4. Adding Your Translation:', ''],
        ['   - Option 1: Add a new column (D, E, etc.) for each target language', ''],
        ['   - Option 2: Create a copy of this file for each language', ''],
        ['   - Include language code in column header (e.g., "Spanish (es)")', ''],
        ['', ''],
        ['5. Quality Assurance:', ''],
        ['   - Review translations in context of their screen', ''],
        ['   - Check that button labels fit on buttons', ''],
        ['   - Ensure error messages make sense', ''],
        ['   - Test that placeholders are preserved correctly', ''],
        ['', ''],
        ['Questions or Issues?', ''],
        ['Contact the development team for clarification on any strings.', ''],
    ]

    df = pd.DataFrame(instructions)
    df.to_excel(writer, sheet_name='Instructions', index=False, header=False)

def create_summary_sheet(writer, sheets_data, metadata):
    """Create summary sheet with statistics"""
    summary = [
        ['StepzSync Translation Summary', ''],
        ['', ''],
        ['Extraction Date:', metadata['extractedAt'][:10]],
        ['Total Strings Found:', metadata['totalStrings']],
        ['Files Processed:', metadata['filesProcessed']],
        ['', ''],
        ['Strings by Category:', 'Count'],
    ]

    # Add category counts
    for category, strings in sheets_data.items():
        summary.append([category, len(strings)])

    summary.extend([
        ['', ''],
        ['Translation Progress:', ''],
        ['This section can be updated as translation progresses', ''],
        ['', ''],
        ['Category', 'Total', 'Translated', 'Progress %'],
    ])

    # Add progress tracking rows
    for category, strings in sheets_data.items():
        summary.append([category, len(strings), 0, '0%'])

    df = pd.DataFrame(summary)
    df.to_excel(writer, sheet_name='Summary', index=False, header=False)

def apply_formatting(filename, sheets_data):
    """Apply professional formatting to the Excel file"""
    wb = load_workbook(filename)

    # Define styles
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')

    instructions_title_fill = PatternFill(start_color='2E75B6', end_color='2E75B6', fill_type='solid')
    instructions_title_font = Font(name='Calibri', size=16, bold=True, color='FFFFFF')

    section_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
    section_font = Font(name='Calibri', size=11, bold=True)

    cell_alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Format Instructions sheet
    if 'Instructions' in wb.sheetnames:
        ws = wb['Instructions']
        ws.column_dimensions['A'].width = 60
        ws.column_dimensions['B'].width = 30

        # Title
        ws['A1'].fill = instructions_title_fill
        ws['A1'].font = instructions_title_font
        ws.merge_cells('A1:B1')

        # Section headers
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
            cell = row[0]
            if cell.value and isinstance(cell.value, str):
                if cell.value.endswith(':') and len(cell.value) < 50:
                    cell.fill = section_fill
                    cell.font = section_font

    # Format Summary sheet
    if 'Summary' in wb.sheetnames:
        ws = wb['Summary']
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 15

        ws['A1'].fill = instructions_title_fill
        ws['A1'].font = instructions_title_font
        ws.merge_cells('A1:B1')

        # Format header rows
        for row in ws.iter_rows(min_row=7, max_row=7):
            for cell in row:
                cell.fill = header_fill
                cell.font = header_font

    # Format category sheets
    for sheet_name in sheets_data.keys():
        sheet_display_name = sheet_name[:31]
        if sheet_display_name in wb.sheetnames:
            ws = wb[sheet_display_name]

            # Set column widths
            ws.column_dimensions['A'].width = 60  # English Text
            ws.column_dimensions['B'].width = 40  # Screen/Context
            ws.column_dimensions['C'].width = 50  # Notes

            # Format header row
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border

            # Format data rows
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                for cell in row:
                    cell.alignment = cell_alignment
                    cell.border = border

            # Freeze header row
            ws.freeze_panes = 'A2'

            # Add auto-filter
            ws.auto_filter.ref = ws.dimensions

    wb.save(filename)

def main():
    print('📊 Generating Excel Translation Workbook...\n')

    # Load data
    print('📖 Loading extracted strings...')
    data = load_strings_data()

    print(f'   Found {len(data["strings"])} strings from {data["metadata"]["filesProcessed"]} files')

    # Clean and deduplicate
    print('🧹 Cleaning and filtering strings...')
    cleaned_strings = clean_and_deduplicate_strings(data['strings'])
    print(f'   After filtering: {len(cleaned_strings)} user-facing strings')

    # Organize into sheets
    print('📑 Organizing into categories...')
    sheets_data = organize_by_sheets(cleaned_strings)

    # Update metadata with cleaned count
    data['metadata']['totalStrings'] = len(cleaned_strings)

    # Print category summary
    print('\n📈 Strings per category:')
    for category, strings in sheets_data.items():
        if strings:
            print(f'   {category}: {len(strings)} strings')

    # Create Excel
    print('\n📝 Creating Excel workbook...')
    output_file = create_excel_workbook(sheets_data, data['metadata'])

    print(f'\n✅ Translation workbook created successfully!')
    print(f'📄 File: {output_file}')
    print(f'📊 Total translatable strings: {len(cleaned_strings)}')
    print(f'📋 Number of sheets: {len([s for s in sheets_data.values() if s])} category sheets + 2 info sheets')

if __name__ == '__main__':
    main()
